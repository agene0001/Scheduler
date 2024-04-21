import os
import uuid

from flaskr.auth import login_required
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import date

from werkzeug.utils import secure_filename
from xlsxwriter import Workbook
from flask import (
    Blueprint, flash, g, redirect, render_template, request, session, url_for, jsonify, json, Response, current_app,
    send_file
)
# from flaskr.db import get_db
from openpyxl import Workbook
from flaskr import db
from flaskr.models import Charts, TimeCards

wb = Workbook()
bp = Blueprint('main_page', __name__)


@bp.route("/clockin-clockout")
@login_required
def clockin():
    return render_template("scheduler/clock_in_out.html")


def getTimecards(user_id):
    # event1 =db.collection('events').get()
    # print(event1[0].to_dict())
    events = db.collection('timeCards').where("timeId", "==", user_id).order_by('data.CIDay').stream()
    return events


@bp.route("/clocksubmitpage", methods=('GET', 'POST'))
@login_required
def clocksubmit():
    if request.method == "POST":
        if session['CICO'] != request.form['ci-co']:
            time1 = request.form['time'].split('at')
            time = time1[1]
            date = time1[0].strip()
            day = int("".join(date.split("/")))
            CI_CO = request.form['ci-co']
            name = session['name'].lower()
            house = request.form['house'].lower()
            session['CICO'] = request.form['ci-co']

            # tups = db.execute('SELECT * FROM user WHERE id=?', (session.get('user_id'),)).fetchall()
            if session['CICO'] == 'CI':
                card = TimeCards(name=name, house=house, dayNum=day, CIDay=date,
                                 CItime=time.strip(), timeId=session.get('user_id'))
                db.collection('timeCards').document().set(card.to_dict())
                # db.execute(
                #     'INSERT INTO timeCards (time_id,name,house,dayNum,CICO,timeIndex,CIDay,CITime) VALUES (?,?,?,?,?,?,?,?)',
                #     (session.get('user_id'), name, house, day, CI_CO, ind, date, time.strip()))
                # 'cards(id, ind, COday, dayNum, COtime) VALUES(?, ?, ?, ?, ?, ?)'(
            else:
                docs = getTimecards(session.get('user_id'))  # for a better understanding check PEP 448
                # print(last)
                for doc in docs:
                    print(doc.to_dict())
                    docMax = doc

                docMax.reference.update({'data.CODay': date, 'data.COtime': time})
                # doc.update({'CODay':date,'COtime':time.strip()})
                # db.collection('timeCards').document().update({'name':name})
                # db.execute('UPDATE timeCards SET CICO=?, CODay=?, COtime=? WHERE time_id=? AND timeIndex=?',
                #            (CI_CO, date, time.strip(), session.get('user_id'), ind))
                # db.execute('UPDATE user SET timeIndex=?, CICO=? WHERE id=?', (ind + 1, CI_CO, session.get('user_id')))
            # db.commit()
            return render_template("scheduler/clocksubmit.html", time=time, CICO=CI_CO, name=name, house=house)
        else:
            if request.form['ci-co'] == 'CI':
                flash("Can't Clock you in as you are already clocked in")
            else:
                flash("Can't Clock you out as you are already clocked out")
            return redirect(request.url)
    return render_template('scheduler/clock_in_out.html')


@bp.route('/timesheet')
@login_required
def timesheet():
    # db = get_db()
    timesheets = db.collection('timeCards').where('timeId', '==', session.get('user_id')).order_by('timeId').stream()
    timesheets = [doc.to_dict()['data'] for doc in timesheets]
    return render_template('scheduler/timesheet.html', dict=timesheets)


def landing():
    return render_template("scheduler/landing.html")


def landingA():
    days = ['Sunday', 'Saturday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    for i in days:
        if i in session:
            session.pop(i)
    return render_template("admin/landingA.html")


@bp.route("/home")
@bp.route('/landingA')
@bp.route('/landing')
@login_required
def home():
    if session['admin']:
        return landingA()
    else:
        return landing()


@bp.route("/schedule", methods=('GET', 'POST'))
@login_required
def schedule():
    if request.method == 'POST':
        # db = get_db()
        if request.form.get('delete') == 'True':
            name = request.form.get('delete.name').lower().strip()
            CIDate = request.form.get('delete.CIDate').strip()
            CODate = request.form.get('delete.CODate').strip()
            CITime = request.form.get('delete.CITime').strip()
            timeId = request.form.get('delete.timeId').strip()
            results = db.collection('timeCards').where("name", "==", name).stream()
            # results = db.execute('SELECT * FROM timeCards WHERE name = ? AND CIDay = ? AND CODay = ? AND CItime = ?',
            #                      (name, CIDate, CODate, CITime)).fetchall()
            if not results:
                flash('Invalid Timecard')
                return redirect(request.url)
            else:
                for result in results:
                    data = result.to_dict()

                    if data['name'] == name and data['data']['CIDay'] == CIDate and data['data']['CODay'] == CODate and data['data']['CItime'] == CITime:
                        result.reference.delete()
                # db.execute('DELETE FROM timeCards WHERE name = ? AND CIDay = ? AND CODay = ? AND CItime = ?',
                #            (name, CIDate, CODate, CITime))
                # db.commit()
                flash('Complete')
                return render_template('admin/schedules.html')
        # elif request.form.get('add') == 'True':
        #     name = request.form.get('add.name').strip()
        #     CIDate = request.form.get('add.CIDate').strip()
        #     CODate = request.form.get('add.CODate').strip()
        #     CITime = request.form.get('add.CITime').strip()
        #     COTime = request.form.get('add.COTime').strip()
        #     house = request.form.get('add.house').strip().lower()
        #
        #     date = CIDate.split('-')
        #     COdate = CODate.split('-')
        #     temp = date[0]
        #     temp1 = COdate[0]
        #     COdate[0] = COdate[1]
        #     COdate[1] = COdate[2]
        #     COdate[2] = temp1
        #     date[0] = date[1]
        #     date[1] = date[2]
        #     date[2] = temp
        #     CIDate = '/'.join(date)
        #     CODate = '/'.join(COdate)
        #     daynum = int(''.join(date))
        #     card = TimeCards(name=name, house=house, dayNum=daynum, CIDay=CIDate,CODay=CODate, CItime=CITime, timeId=session.get('user_id'))
        #     db.execute(
        #         'INSERT INTO timeCards(name,CIDay,CODay,CItime,COtime,dayNum,house,CICO) VALUES (?,?,?,?,?,?,?,?)',
        #         (name, CIDate, CODate, CITime, COTime, daynum, house, 'CO'))
        #     db.commit()
        #     flash('Complete')
        #     return render_template('admin/schedules.html')
        sortBy = request.form['sort']
        val = request.form['value']
        days = request.form['filter'].split("-")
        first = days[0]
        days[0] = days[1]
        days[1] = days[2]
        days[2] = first
        dayNum = int("".join(days))
        tups = ()
        match sortBy:
            case 'House':
                house = val.lower()
                docs = db.collection('timeCards').get()
                for doc in docs:
                    docData = doc.to_dict()
                    if docData['data']['house'] == house and docData['data']['dayNum'] >= dayNum:
                        tups += (docData['data'] | {'name': docData['name']},)
                # tups = db.execute('SELECT * FROM timeCards WHERE house=? AND dayNum>=? AND CICO=? ORDER BY time_id',
                #                   (val, dayNum, 'CO')).fetchall()

                # tups = db.execute('SELECT * FROM timeCards WHERE day=? AND dayNum>=? AND CICO=? ORDER BY name',
                #                   (val, dayNum, 'CO')).fetchall()
            case 'Name':
                val = val.lower()
                names = val.split(', ')
                docs = db.collection('timeCards').get()
                for doc in docs:
                    data = doc.to_dict()
                    if data['name'] in names and data['data']['dayNum'] >= dayNum:
                        tups += (data['data'] | {'name': data['name']},)
                # str = 'SELECT * FROM timeCards WHERE '
                # tup = (dayNum,)
                # for i in range(len(names)):
                #     tup = (names[i],) + tup
                #     if i == 0:
                #         str += '(name LIKE ? '
                #     else:
                #         str += 'OR name LIKE ? '
                # tup += ("CO",)
                # str += ') AND dayNum>=? AND CICO=? ORDER BY name'
                # tups = db.execute(str,
                #                   tup).fetchall()
            case 'Everything':
                tups = db.collection('timeCards').get()
                tups = list(map(lambda card: card.to_dict(), tups))
                lis = []
                for i in tups:
                    # print(i['data']['dayNum'])
                    if i['data']['dayNum'] >= dayNum:
                        lis.append(i['data'] | {'name': i['name']})
                tups = lis
                # tups = db.execute('SELECT * FROM timeCards WHERE dayNum>= ? AND CICO=? ORDER BY name',
                #                   (dayNum, 'CO')).fetchall()
        return render_template("admin/schedules.html", vals=tups)
    return render_template("admin/schedules.html")


@bp.route('/charts', methods=('GET', 'POST'))
@login_required
def charts():
    print(request.form)
    # this line prints out the form to the browser
    print(jsonify(request.form.to_dict()))
    return render_template("admin/chartsList.html")


@bp.route('/download/<filename>')
@login_required
def download_file(filename):
    # Serve a file for download
    return send_file(os.path.join(current_app.instance_path, current_app.config['UPLOAD_FOLDER'], filename),
                     as_attachment=True)


@bp.route('/files', methods=('GET', 'POST'))
@login_required
def list_files():
    if request.method == 'POST':
        if request.form.get('delete') is not None:
            if os.path.exists(current_app.instance_path + '\\' + os.path.join(current_app.config['UPLOAD_FOLDER'],
                                                                              request.form['delete'])):
                os.remove(current_app.instance_path + '\\' + os.path.join(current_app.config['UPLOAD_FOLDER'],
                                                                          request.form['delete']))
                flash('Successful')
                files = os.listdir(current_app.instance_path + '\\' + current_app.config['UPLOAD_FOLDER'])
                return render_template('files.html', files=files)
            else:
                flash('File doesnt exist')
                return redirect(request.url)

        file = request.files['upload']
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        if file:
            filename = secure_filename(file.filename)
            file.save(current_app.instance_path + '\\' + os.path.join(current_app.config['UPLOAD_FOLDER'], filename))
    # Get a list of files in the uploads folder
    files = os.listdir(current_app.instance_path + '\\' + current_app.config['UPLOAD_FOLDER'])
    return render_template('files.html', files=files)


@bp.route('/homeAidChart', methods=('GET', 'POST'))
@login_required
def homeAidChart():
    if request.method == 'POST':
        # print(request.form)
        # this line prints out the form to the browser

        # print(jsonify(request.form.to_dict()))
        roomNum = request.form.getlist('roomNum')[0]
        session['days'] = request.form.getlist('day')
        session['initials'] = request.form.getlist('initials')[0]
        session['clientName'] = request.form.getlist('clientName')[0]
        if roomNum != '':
            session['roomNum'] = roomNum
        return render_template("admin/healthAidChartDays.html", ind=0)
    else:
        return render_template("admin/homeHealthAidChart.html")


@bp.route('/chartComplete', methods=('GET', 'POST'))
@login_required
def chartComplete():
    days = {}
    dayStr = ['Sunday', 'Saturday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday']
    # if you ever want to redirect using same method used
    # return redirect(url_for('main_page.charts'), code=307)
    for i in dayStr:
        if i in session:
            # print(session[i])
            days[i] = session[i]
        else:
            days[i] = {}
    name = "-".join((session['clientName'].split(', ')))
    uni = uuid.uuid4().hex[:3]
    wb = createWorkbooks(f'Miles_vents_healthcare_Chart-{date.today()}-{name}{uni}.xlsx')
    wb.makeHealthChart()
    wb.addVals(days, session['initials'], session['clientName'])
    wb.save()
    flash('Chart Complete')
    return redirect(url_for('main_page.homeAidChart'))


@bp.route('/homeAidChartDays', methods=('GET', 'POST'))
@login_required
def homeAidChartDays():
    if request.method == 'POST':
        # currDay = day(session['days'][int(request.form['index'])], request.form.getlist('dressing'),
        #               request.form.getlist('hygiene'), request.form.getlist('toilet'), request.form.getlist('activity'),
        #               request.form.getlist('homemaking'),request.form.getlist('medicalAssistance'))
        # this line prints out the form to the browser
        # print(jsonify(request.form.to_dict()))
        day = f"{session['days'][int(request.form['index'])]}"
        session[day] = {'dressing': request.form.getlist('dressing'), 'hygiene': request.form.getlist('hygiene'),
                        'toilet': request.form.getlist('toilet'), 'activity': request.form.getlist('activity'),
                        'homemaking': request.form.getlist('homemaking'),
                        'medicalAssistance': request.form.getlist('medicalAssistance')}
        if len(session['days']) > int(request.form['index']) + 1:
            return render_template("admin/healthAidChartDays.html", ind=int(request.form['index']) + 1)
        else:
            days = {}
            session['days'] = None

            return redirect(url_for("main_page.chartComplete"))
    else:
        return charts()


class createWorkbooks():
    def __init__(self, name='Miles_vents_healthcare_Chart'):
        self.wb = Workbook()
        self.ws = wb.active
        self.name = name

    def assignCell(self, cell, value):
        self.ws[cell] = value

    def getCell(self, cell):
        return self.ws[cell]

    def makeHealthChart(self):
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for col in self.ws.iter_cols(min_row=6, min_col=4, max_col=10, max_row=51):
            for cell in col:
                cell.border = thin_border

        wb = Workbook()
        ws = wb.get_active_sheet()
        # property cell.border should be used instead of cell.style.border
        ws.cell(row=3, column=2).border = thin_border
        # bolds
        bolds = ['C4', 'D5', 'E5', 'F5', 'G5', 'H5', 'I5', 'J5', 'C6', 'C11', 'C21', 'C29']
        for i in bolds:
            self.ws[i].font = Font(bold=True)
            self.ws[i].alignment = Alignment(horizontal='left')
        self.ws['C4'] = 'Start Date:'
        self.ws['D5'] = 'Saturday'
        self.ws['E5'] = 'Sunday'
        self.ws['F5'] = 'Monday'
        self.ws['G5'] = 'Tuesday'
        self.ws['H5'] = 'Wednesday'
        self.ws['I5'] = 'Thursday'
        self.ws['J5'] = 'Friday'
        self.ws['C6'] = 'Dressing:'
        self.ws['C11'] = 'Hygiene:'
        self.ws['C21'] = 'Toileting:'
        self.ws['C29'] = 'Exercise:'

        # bolds right-align
        boldsR = ['C25', 'C44', 'C50', 'C33', 'C45', 'C51']

        for i in boldsR:
            self.ws[i].font = Font(bold=True)
            self.ws[i].alignment = Alignment(horizontal='right')

        self.ws['C33'] = 'Homemaking:'
        self.ws['C45'] = 'Vital Signs:'
        self.ws['C51'] = 'Treatments:'
        self.ws['C25'] = 'Meal/Plate Prep:'
        self.ws['C44'] = 'Medication Assistance/MAR:'
        self.ws['C50'] = 'Behavior/Orientation:'

        # title
        self.ws['B1'] = 'Miles Vents INC Assisted Living Charting Sheet'
        self.ws['B1'].font = Font(size=38)
        # subheading
        subHeads = ['B3', 'L6', 'L9', 'L10']
        self.ws['B3'] = ('Initial activities completed. The circled initials indicate '
                         'that the resident did not receive the intervention. Documented reasons beside')
        self.ws['L6'] = 'Reasons:'
        self.ws['L9'] = 'Residents:'
        for i in subHeads:
            self.ws[i].font = Font(size=15)

        #         values
        values = ['C7', 'C8', 'C9', 'C10', 'C12', 'C13', 'C14', 'C15', 'C16', 'C17', 'C18', 'C19', 'C20', 'C22', 'C23',
                  'C24', 'C26', 'C27', 'C28', 'C30', 'C31', 'C32', 'C34', 'C35', 'C36', 'C37', 'C38', 'C39', 'C40',
                  'C41', 'C42', 'C43', 'C46', 'C47', 'C48', 'C49']
        for i in values:
            self.ws[i].alignment = Alignment(horizontal='right')
        # dressing
        self.ws['C7'] = 'Clothes'
        self.ws['C8'] = 'Hearing Aid'
        self.ws['C9'] = 'Elastic Stockings/TEDs'
        self.ws['C10'] = 'Braces/Orthotics'
        # Hygiene
        self.ws['C12'] = 'Bath'
        self.ws['C13'] = 'Shampoo'
        self.ws['C14'] = 'Oral Care'
        self.ws['C15'] = 'Hair Care'
        self.ws['C16'] = 'Dentures'
        self.ws['C17'] = 'Shaving'
        self.ws['C18'] = 'Nail Care'
        self.ws['C19'] = 'Foot Care'
        self.ws['C20'] = 'Skin Care'

        #         Toileting
        self.ws['C22'] = 'Toileting Assistance'
        self.ws['C23'] = 'Incontinent Pads'
        self.ws['C24'] = 'Bowel Movements'
        # meal prep
        self.ws['C26'] = 'AM'
        self.ws['C27'] = 'Noon'
        self.ws['C28'] = 'PM'
        # exercise
        self.ws['C30'] = 'Excercise Reminder'
        self.ws['C31'] = 'Standby/Ambulation'
        self.ws['C32'] = 'Medication Reminder'
        # Homemaking
        self.ws['C34'] = 'Clean Kitchen'
        self.ws['C35'] = 'Remove Garbage'
        self.ws['C36'] = 'Monitor Foods'
        self.ws['C37'] = 'Shopping'
        self.ws['C38'] = 'Clean Bathroom'
        self.ws['C39'] = 'Clean Bedroom'
        self.ws['C40'] = 'Change Linens'
        self.ws['C41'] = 'Monitor Clothing'
        self.ws['C42'] = 'Laundry'
        self.ws['C43'] = 'Vacuum/Dust'

        # medication assistance
        self.ws['C46'] = 'TPR'
        self.ws['C47'] = 'BP'
        self.ws['C48'] = 'Weight'
        self.ws['C49'] = 'Blood Glucose'

    def addVals(self, dict, initials, clientName):
        for i in dict.items():
            if i[1] != {}:
                match i[0]:
                    case 'Saturday':
                        letter = 'D'
                        self.addValHelp(letter, i[1], initials)

                    case 'Sunday':
                        letter = 'E'
                        self.addValHelp(letter, i[1], initials)

                    case 'Monday':
                        letter = 'F'
                        self.addValHelp(letter, i[1], initials)

                    case 'Tuesday':
                        letter = 'G'
                        self.addValHelp(letter, i[1], initials)

                    case 'Wednesday':
                        letter = 'H'
                        self.addValHelp(letter, i[1], initials)

                    case 'Thursday':
                        letter = 'I'
                        self.addValHelp(letter, i[1], initials)

                    case 'Friday':
                        letter = 'J'
                        self.addValHelp(letter, i[1], initials)
        clientName = clientName.strip(', ')
        self.ws['L10'] = clientName

    def addValHelp(self, letter, dict, initials):
        for i in dict.items():
            for i in i[1]:
                match i:
                    case 'clothes':
                        self.ws[letter + '7'] = initials.upper()

                    case 'hearingAid':
                        self.ws[letter + '8'] = initials.upper()

                    case 'stockings':
                        self.ws[letter + '9'] = initials.upper()

                    case 'braces':
                        self.ws[letter + '10'] = initials.upper()

                    case 'bath':
                        self.ws[letter + '12'] = initials.upper()

                    case 'shampoo':
                        self.ws[letter + '13'] = initials.upper()

                    case 'oral':
                        self.ws[letter + '14'] = initials.upper()

                    case 'hair':
                        self.ws[letter + '15'] = initials.upper()

                    case 'dentures':
                        self.ws[letter + '16'] = initials.upper()

                    case 'shaving':
                        self.ws[letter + '17'] = initials.upper()

                    case 'skincare':
                        self.ws[letter + '20'] = initials.upper()

                    case 'nailcare':
                        self.ws[letter + '18'] = initials.upper()

                    case 'footcare':
                        self.ws[letter + '19'] = initials.upper()

                    case 'footcare':
                        self.ws[letter + '19'] = initials.upper()

                    case 'toiletassistance':
                        self.ws[letter + '22'] = initials.upper()

                    case 'pads':
                        self.ws[letter + '23'] = initials.upper()

                    case 'bowel':
                        self.ws[letter + '24'] = initials.upper()

                    case 'AM':
                        self.ws[letter + '26'] = initials.upper()
                        self.ws[letter + '25'] = initials.upper()

                    case 'PM':
                        self.ws[letter + '28'] = initials.upper()
                        self.ws[letter + '25'] = initials.upper()

                    case 'Noon':
                        self.ws[letter + '27'] = initials.upper()
                        self.ws[letter + '25'] = initials.upper()

                    case 'AM/PM':
                        self.ws[letter + '26'] = initials.upper()
                        self.ws[letter + '28'] = initials.upper()
                        self.ws[letter + '25'] = initials.upper()

                    case 'standby':
                        self.ws[letter + '31'] = initials.upper()

                    case 'medReminder':
                        self.ws[letter + '32'] = initials.upper()

                    case 'exerciseReminder':
                        self.ws[letter + '30'] = initials.upper()

                    case 'kitchen':
                        self.ws[letter + '34'] = initials.upper()

                    case 'garbage':
                        self.ws[letter + '35'] = initials.upper()

                    case 'food':
                        self.ws[letter + '36'] = initials.upper()

                    case 'shopping':
                        self.ws[letter + '37'] = initials.upper()

                    case 'bathroom':
                        self.ws[letter + '38'] = initials.upper()

                    case 'bedroom':
                        self.ws[letter + '39'] = initials.upper()

                    case 'linen':
                        self.ws[letter + '40'] = initials.upper()

                    case 'clothing':
                        self.ws[letter + '41'] = initials.upper()

                    case 'laundry':
                        self.ws[letter + '42'] = initials.upper()

                    case 'vacuum':
                        self.ws[letter + '43'] = initials.upper()

                    case 'companion':
                        self.ws[letter + '51'] = initials.upper()

                    case 'MAR':
                        self.ws[letter + '44'] = initials.upper()

                    case 'TPR':
                        self.ws[letter + '46'] = initials.upper()

                    case 'BP':
                        self.ws[letter + '47'] = initials.upper()

                    case 'Weight':
                        self.ws[letter + '48'] = initials.upper()

                    case 'BloodGlucose':
                        self.ws[letter + '49'] = initials.upper()

                    case 'Behavior/Orientation':
                        self.ws[letter + '50'] = initials.upper()

    def save(self):
        wb.save(current_app.instance_path + '\\' + os.path.join(current_app.config['UPLOAD_FOLDER'], self.name))
