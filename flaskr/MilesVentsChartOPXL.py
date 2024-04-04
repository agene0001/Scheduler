from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


#
# font = Font(name='Calibri',
#                      size=11,
#                      bold=False,
#                      italic=False,
#                      vertAlign=None,
#                      underline='none',
#                      strike=False,
#                      color='FF000000')
#      fill = PatternFill(fill_type=None,
#                      start_color='FFFFFFFF',
#                      end_color='FF000000')
#      border = Border(left=Side(border_style=None,
#                                color='FF000000'),
#                      right=Side(border_style=None,
#                                 color='FF000000'),
#                      top=Side(border_style=None,
#                               color='FF000000'),
#                      bottom=Side(border_style=None,
#                                  color='FF000000'),
#                      diagonal=Side(border_style=None,
#                                    color='FF000000'),
#                      diagonal_direction=0,
#                      outline=Side(border_style=None,
#                                   color='FF000000'),
#                      vertical=Side(border_style=None,
#                                    color='FF000000'),
#                      horizontal=Side(border_style=None,
#                                     color='FF000000')
#                     )
#      alignment=Alignment(horizontal='general',
#                          vertical='bottom',
#                          text_rotation=0,
#                          wrap_text=False,
#                          shrink_to_fit=False,
#                          indent=0)
#      number_format = 'General'
#      protection = Protection(locked=True,
#                              hidden=False)
#

class createWorkbooks():
    def __init__(self, name=None):
        self.wb = Workbook();
        self.ws = self.wb.active()
        self.name = name

    def assignCell(self, cell, value):
        self.ws[cell] = value

    def getCell(self, cell):
        return self.ws[cell]

    def makeHealthChart(self):
        # bolds
        bolds = ['C4', 'D5', 'E5', 'F5', 'G5', 'H5', 'I5', 'J5', 'C6', 'C11', 'C21', 'C29']
        for i in bolds:
            self.ws[i].font = Font(bold=True)
            self.ws[i].alignment = Alignment(horizontal='left')
        self.ws['C4'] = 'Start Date:'
        self.ws['D5'] = 'Saturday'
        self.ws['E5'] = 'Sunday'
        self.ws['F5'] = 'Monday'
        self.ws['G5'] = 'Tuesday'
        self.ws['H5'] = 'Wednesday'
        self.ws['I5'] = 'Thursday'
        self.ws['J5'] = 'Friday'
        self.ws['C6'] = 'Dressing:'
        self.ws['C11'] = 'Hygiene:'
        self.ws['C21'] = 'Toileting:'
        self.ws['C29'] = 'Exercise:'

        # bolds right-align
        boldsR = ['C25', 'C44', 'C50', 'C33', 'C45', 'C51']

        for i in boldsR:
            self.ws[i].font = Font(bold=True)
            self.ws[i].alignment = Alignment(horizontal='right')

        self.ws['C33'] = 'Homemaking:'
        self.ws['C45'] = 'Vital Signs:'
        self.ws['C51'] = 'Treatments:'
        self.ws['C25'] = 'Meal/Plate Prep:'
        self.ws['C44'] = 'Medication Assistance/MAR:'
        self.ws['C50'] = 'Behavior/Orientation:'

        # title
        self.ws['B1'] = 'Miles Vents INC Assisted Living Charting Sheet'
        self.ws['B1'].font = Font(size=38)
        # subheading
        subHeads = ['B3', 'L6']
        self.ws['B3'] = ('Initial activities completed. The circled initials indicate '
                         'that the resident did not receive the intervention. Documented reasons beside')
        self.ws['L6'] = 'Reasons:'
        for i in subHeads:
            self.ws[i].font = Font(size=15)

        #         values
        values = ['C7', 'C8', 'C9', 'C10', 'C12', 'C13', 'C14', 'C15', 'C16', 'C17', 'C18', 'C19', 'C20', 'C22', 'C23',
                  'C24', 'C26', 'C27', 'C28', 'C30', 'C31', 'C32', 'C34', 'C35', 'C36', 'C37', 'C38', 'C39', 'C40',
                  'C41', 'C42', 'C43', 'C46', 'C47', 'C48', 'C49']
        for i in values:
            self.ws[i].alignment = Alignment(horizontal='right')
        # dressing
        self.ws['C7'] = 'Clothes'
        self.ws['C8'] = 'Hearing Aid'
        self.ws['C9'] = 'Elastic Stockings/TEDs'
        self.ws['C10'] = 'Braces/Orthotics'
        # Hygiene
        self.ws['C12'] = 'Bath'
        self.ws['C13'] = 'Shampoo'
        self.ws['C14'] = 'Oral Care'
        self.ws['C15'] = 'Hair Care'
        self.ws['C16'] = 'Dentures'
        self.ws['C17'] = 'Shaving'
        self.ws['C18'] = 'Nail Care'
        self.ws['C19'] = 'Foot Care'
        self.ws['C20'] = 'Skin Care'

        #         Toileting
        self.ws['C22'] = 'Toileting Assistance'
        self.ws['C23'] = 'Incontinent Pads'
        self.ws['C24'] = 'Bowel Movements'
        # meal prep
        self.ws['C26'] = 'AM'
        self.ws['C27'] = 'Noon'
        self.ws['C28'] = 'PM'
        # exercise
        self.ws['C30'] = 'Prescribed'
        self.ws['C31'] = 'Ambulation'
        self.ws['C32'] = 'Transfer Assistance'
        # Homemaking
        self.ws['C34'] = 'Clean Kitchen'
        self.ws['C35'] = 'Remove Garbage'
        self.ws['C36'] = 'Monitor Foods'
        self.ws['C37'] = 'Shopping'
        self.ws['C38'] = 'Clean Bathroom'
        self.ws['C39'] = 'Clean Bedroom'
        self.ws['C40'] = 'Change Linens'
        self.ws['C41'] = 'Monitor Clothing'
        self.ws['C42'] = 'Laundry'
        self.ws['C43'] = 'Vacuum/Dust'

        # medication assistance
        self.ws['C46'] = 'TPR'
        self.ws['C47'] = 'BP'
        self.ws['C48'] = 'Weight'
        self.ws['C49'] = 'Blood Glucose'

    def addVals(self,dict,initials):
        for i in dict.items():
            match i[0]:
                case 'Saturday':
                    for j in i[1].items():
                       print(j)
                    
                case 'Sunday':
                    for j in i[1].items():
                       print(j)
                    
                case 'Monday':
                    for j in i[1].items():
                       print(j)
                    
                case 'Tuesday':
                    for j in i[1].items():
                       print(j)
                    
                case 'Wednesday':
                    for j in i[1].items():
                       print(j)
                    
                case 'Thursday':
                    for j in i[1].items():
                       print(j)
                    
                case 'Friday':
                    for j in i[1].items():
                       print(j)
                    
    def addValHelp(self,letter):
        pass
    def save(self):
        if self.name != None:
            self.wb.save(self.name)
        else:
            self.wb.save("Miles_Vents_Chart.xlsx")


# if __name__ == '__main__':
#     wb = Workbook()
#     ws = wb.active
#     wbObj = createWorkbooks(ws)
#     wbObj.makeHealthChart()
#     wbObj.save()
